from datetime import timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from loguru import logger

from sqlalchemy.orm import aliased
from src.models import User, EventFeedback, Event, ExpertRating

from src.models import Event, EventFeedback, ExpertProfile


async def generate_event_excel_report(db: AsyncSession, event_id: int) -> str | None:
    logger.info(f"Starting Excel report generation for event_id: {event_id}")

    event_query = (
        select(Event)
        .options(selectinload(Event.expert).selectinload(ExpertProfile.user))
        .where(Event.id == event_id)
    )

    event_res = await db.execute(event_query)
    event = event_res.scalars().first()

    if not event:
        logger.warning(f"Event {event_id} not found.")
        return None

    # 2. Получаем отзывы
    feedbacks_query = (
        select(EventFeedback)
        .where(EventFeedback.event_id == event_id)
        .order_by(EventFeedback.created_at.desc())
    )
    feedbacks_res = await db.execute(feedbacks_query)
    feedbacks = feedbacks_res.scalars().all()

    # 3. Расчет времени
    msk_tz = timezone(timedelta(hours=3))
    start_dt = event.event_date.astimezone(msk_tz)
    end_dt = start_dt + timedelta(minutes=event.duration_minutes)

    date_str = f"{start_dt.strftime('%d.%m.%Y %H:%M')} — {end_dt.strftime('%H:%M')}"

    # 4. Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по мероприятию"

    # --- ШАПКА МЕРОПРИЯТИЯ ---
    ws["A1"] = "Название мероприятия:"
    ws["B1"] = event.name
    ws["A1"].font = Font(bold=True)

    ws["A2"] = "Промо-слово:"
    ws["B2"] = event.promo_word
    ws["A2"].font = Font(bold=True)

    ws["A3"] = "Дата проведения:"
    ws["B3"] = date_str
    ws["A3"].font = Font(bold=True)

    ws["A4"] = "Всего голосов:"
    ws["B4"] = len(feedbacks)
    ws["A4"].font = Font(bold=True)

    # Отступ
    ws.append([])

    # --- ТАБЛИЦА ГОЛОСОВ ---
    headers = ["Дата и время", "VK ID Слушателя", "Голос", "Комментарий"]
    ws.append(headers)

    # Стилизация заголовка таблицы
    header_fill = PatternFill(
        start_color="4A76A8", end_color="4A76A8", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, cell in enumerate(ws[6], 1):  # 6-я строка - заголовки
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Заполнение данными
    for fb in feedbacks:
        vote_text = "Нейтрально"
        if fb.rating_snapshot == 1:
            vote_text = "Доверяю (+1)"
        elif fb.rating_snapshot == -1:
            vote_text = "Не доверяю (-1)"

        vote_dt = (
            fb.created_at.astimezone(msk_tz).strftime("%d.%m.%Y %H:%M:%S")
            if fb.created_at
            else "-"
        )
        comment = fb.comment or ""
        voter_link = f"https://vk.com/id{fb.voter_id}"

        ws.append([vote_dt, voter_link, vote_text, comment])

    # Настройка ширины колонок
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 50

    # Сохранение файла
    filename = f"report_event_{event.promo_word}_{event_id}.xlsx"
    file_path = f"/tmp/{filename}"
    wb.save(file_path)

    logger.success(f"Excel report saved to {file_path}")
    return file_path


async def generate_admin_expert_excel_report(db: AsyncSession, expert_id: int) -> str | None:
    logger.info(f"Generating admin Excel report for expert {expert_id}")

    # 1. Получаем данные эксперта
    expert_res = await db.execute(select(User).where(User.vk_id == expert_id))
    expert = expert_res.scalars().first()
    if not expert:
        return None

    expert_name = f"{expert.first_name} {expert.last_name or ''}".strip()

    # 2. Получаем всю историю (Хронологию)
    Voter = aliased(User)
    query = (
        select(EventFeedback, Voter, Event)
        .join(Voter, EventFeedback.voter_id == Voter.vk_id)
        .outerjoin(Event, EventFeedback.event_id == Event.id)
        .where(EventFeedback.expert_id == expert_id)
        .order_by(EventFeedback.created_at.asc())
    )
    result = await db.execute(query)
    history_rows = result.all()

    # 3. Получаем текущие активные голоса для 2-го листа
    ratings_query = select(ExpertRating).where(ExpertRating.expert_id == expert_id)
    ratings_res = await db.execute(ratings_query)
    active_ratings = {
        (r.voter_id, r.rating_type): r.vote_value
        for r in ratings_res.scalars().all()
    }

    # 4. Настраиваем Excel
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "История действий"
    ws2 = wb.create_sheet(title="Текущий статус")

    headers_1 =[
        "Дата", "Голос", "Тип", "VK ID слушателя", "Имя слушателя",
        "VK ID эксперта", "Имя эксперта", "Название мероприятия",
        "ID мероприятия", "Промослово", "Текст обратной связи"
    ]
    ws1.append(headers_1)

    headers_2 =[
        "Голос", "Тип", "VK ID слушателя", "Имя слушателя",
        "VK ID эксперта", "Имя эксперта", "Название мероприятия",
        "ID мероприятия", "Промослово", "Текст обратной связи"
    ]
    ws2.append(headers_2)

    header_fill = PatternFill(start_color="4A76A8", end_color="4A76A8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    msk_tz = timezone(timedelta(hours=3))
    latest_state = {}

    # 5. Заполняем Лист 1 (Хронология) и собираем данные для Листа 2
    for fb, voter, event in history_rows:
        r_type = "Мероприятие" if fb.event_id else "Народный"
        r_type_key = "expert" if fb.event_id else "community"

        # Обновляем "последнее состояние" для конкретного юзера и типа рейтинга
        latest_state[(voter.vk_id, r_type_key)] = {
            "voter": voter,
            "event": event,
            "comment": fb.comment,
            "r_type_label": r_type
        }

        vote_val = fb.rating_snapshot
        vote_str = f"ОС ({vote_val})" if fb.comment else str(vote_val)

        dt_str = fb.created_at.astimezone(msk_tz).strftime("%d.%m.%Y %H:%M") if fb.created_at else "-"
        voter_name = f"{voter.first_name} {voter.last_name or ''}".strip()
        event_name = event.name if event else "-"
        event_id_str = str(event.id) if event else "-"
        promo = event.promo_word if event else "-"
        comment = fb.comment or ""

        ws1.append([
            dt_str, vote_str, r_type, str(voter.vk_id), voter_name,
            str(expert.vk_id), expert_name, event_name, event_id_str, promo, comment
        ])

    # 6. Заполняем Лист 2 (Фактический текущий статус)
    for (voter_id, r_type_key), data in latest_state.items():
        vote_val = active_ratings.get((voter_id, r_type_key), 0)
        comment = data["comment"] or ""
        vote_str = f"ОС ({vote_val})" if comment else str(vote_val)

        voter = data["voter"]
        event = data["event"]
        voter_name = f"{voter.first_name} {voter.last_name or ''}".strip()
        event_name = event.name if event else "-"
        event_id_str = str(event.id) if event else "-"
        promo = event.promo_word if event else "-"

        ws2.append([
            vote_str, data["r_type_label"], str(voter.vk_id), voter_name,
            str(expert.vk_id), expert_name, event_name, event_id_str, promo, comment
        ])

    # Косметика (ширина колонок)
    for ws in [ws1, ws2]:
        for col_letter in['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col_letter].width = 20

    filename = f"report_admin_expert_{expert_id}.xlsx"
    file_path = f"/tmp/{filename}"
    wb.save(file_path)
    return file_path